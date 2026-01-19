import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
import pandas as pd

csv_path = Path(r"C:\Users\usa\Documents\steward-view-main\data\output\labeled_data.csv")
excel_path = Path(r"C:\Users\usa\Documents\steward-view-main\data\output\labeled_data.xlsx")
charts_dir = Path(r"C:\Users\usa\Documents\steward-view-main\data\output")

print("Current working directory:", Path.cwd())
print("CSV path:", csv_path)
print("Excel output path:", excel_path)

if not csv_path.exists():
    raise FileNotFoundError(f"CSV not found at: {csv_path}")

wb = Workbook()
ws = wb.active
ws.title = "labeled_data"

# If your CSV might not be UTF-8, utf-8-sig is often safer for Windows/Excel exports
with csv_path.open(newline="", encoding="utf-8-sig") as f:
    reader = csv.reader(f)
    row_count = 0
    for row in reader:
        ws.append(row)
        row_count += 1

print("Rows imported:", row_count)

# Create Summary sheet (2nd sheet) and import all summary CSVs
summary_ws = wb.create_sheet("Summary", 1)  # Insert as 2nd sheet (index 1)
current_row = 1

# Import summary.csv (formatted text)
summary_csv_path = charts_dir / 'summary.csv'
if summary_csv_path.exists():
    summary_df = pd.read_csv(summary_csv_path)
    summary_ws['A1'] = 'Summary (Formatted Text)'
    summary_ws['A1'].font = Font(bold=True, size=12)
    current_row = 3
    
    for idx, row in summary_df.iterrows():
        summary_ws[f'A{current_row}'] = row['Summary']
        current_row += 1
    current_row += 2  # Add spacing
    print("Added summary.csv to Summary sheet")

# Import summary_stats.csv
stats_csv_path = charts_dir / 'summary_stats.csv'
if stats_csv_path.exists():
    stats_df = pd.read_csv(stats_csv_path)
    summary_ws[f'A{current_row}'] = 'Summary Statistics'
    summary_ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 2
    
    # Add headers
    summary_ws[f'A{current_row}'] = 'Metric'
    summary_ws[f'B{current_row}'] = 'Value'
    summary_ws[f'A{current_row}'].font = Font(bold=True)
    summary_ws[f'B{current_row}'].font = Font(bold=True)
    current_row += 1
    
    # Add data
    for idx, row in stats_df.iterrows():
        summary_ws[f'A{current_row}'] = row['Metric']
        summary_ws[f'B{current_row}'] = row['Value']
        current_row += 1
    current_row += 2  # Add spacing
    print("Added summary_stats.csv to Summary sheet")

# Adjust column widths
summary_ws.column_dimensions['A'].width = 30
summary_ws.column_dimensions['B'].width = 20

# Create a new sheet for charts
charts_ws = wb.create_sheet("Charts")

# List of chart images to import
chart_files = [
    "spending_by_spender.png",
    "spending_by_category.png",
    "spending_by_vendor.png"
]

# Add charts to the sheet side by side
row_position = 5  # Start from row 5 (leaving 3 empty rows at the top, title at row 4)
# Columns: A, F, M (with spacing between charts to avoid overlap)
columns = ['A', 'F', 'M']

for i, chart_file in enumerate(chart_files):
    chart_path = charts_dir / chart_file
    
    if chart_path.exists():
        # Create image object
        img = Image(chart_path)
        
        # Scale down images to fit nicely (adjust scale as needed)
        # You can adjust these values: 0.25 = 25% of original size
        img.width = int(img.width * 0.25)
        img.height = int(img.height * 0.25)
        
        # Position images side by side using different columns
        col = columns[i] if i < len(columns) else 'A'
        img.anchor = f'{col}{row_position}'
        
        # Add chart title above the image
        title_cell = charts_ws[f'{col}{row_position - 1}']
        title_cell.value = chart_file.replace('.png', '').replace('_', ' ').title()
        from openpyxl.styles import Font
        title_cell.font = Font(bold=True)
        
        # Add image to worksheet
        charts_ws.add_image(img)
        
        print(f"Added chart: {chart_file}")
    else:
        print(f"Warning: Chart not found: {chart_file}")

# Adjust column width for chart titles
charts_ws.column_dimensions['A'].width = 30

# Create 4th sheet with native Excel charts
native_charts_ws = wb.create_sheet("Native Charts")

# Read the labeled data to create aggregated data for charts
labeled_data_df = pd.read_csv(csv_path)
labeled_data_df['amount'] = pd.to_numeric(labeled_data_df['amount'], errors='coerce')

# Prepare aggregated data for charts
spender_totals = labeled_data_df.groupby('spender')['amount'].sum().reset_index()
spender_totals.columns = ['Spender', 'Amount']

category_totals = labeled_data_df.groupby('category')['amount'].sum().reset_index()
category_totals.columns = ['Category', 'Amount']
category_totals = category_totals.sort_values('Amount', ascending=False)

vendor_totals = labeled_data_df.groupby('vendor')['amount'].sum().reset_index()
vendor_totals.columns = ['Vendor', 'Amount']
vendor_totals = vendor_totals.sort_values('Amount', ascending=False)

# Write data for Spending by Spender (columns A-B, starting row 2)
native_charts_ws['A1'] = 'Spending by Spender'
native_charts_ws['A1'].font = Font(bold=True, size=12)
native_charts_ws['A2'] = 'Spender'
native_charts_ws['B2'] = 'Amount'
native_charts_ws['A2'].font = Font(bold=True)
native_charts_ws['B2'].font = Font(bold=True)
for idx, row in spender_totals.iterrows():
    native_charts_ws[f'A{idx+3}'] = row['Spender']
    native_charts_ws[f'B{idx+3}'] = float(row['Amount'])

# Create Pie Chart for Spending by Spender
pie1 = PieChart()
pie1.title = "Spending by Spender"
pie1.title.layout = Layout(manualLayout=ManualLayout(y=0.05, yMode='edge'))  # Position title higher
data = Reference(native_charts_ws, min_col=2, min_row=3, max_row=2+len(spender_totals))  # Start at row 3 to exclude "Amount" header
cats = Reference(native_charts_ws, min_col=1, min_row=3, max_row=2+len(spender_totals))
pie1.add_data(data, titles_from_data=False)
pie1.set_categories(cats)
pie1.legend.position = 'r'  # Position legend to the right
pie1.height = 10
pie1.width = 15  # Increased width to make room for legend on the right
native_charts_ws.add_chart(pie1, "D2")

# Write data for Spending by Category (columns A-B, with 10 rows spacing after row 14)
start_row_category = 25  # 10 rows after row 14, as it was correct before
native_charts_ws[f'A{start_row_category}'] = 'Spending by Category'
native_charts_ws[f'A{start_row_category}'].font = Font(bold=True, size=12)
native_charts_ws[f'A{start_row_category+1}'] = 'Category'
native_charts_ws[f'B{start_row_category+1}'] = 'Amount'
native_charts_ws[f'A{start_row_category+1}'].font = Font(bold=True)
native_charts_ws[f'B{start_row_category+1}'].font = Font(bold=True)
for idx, row in category_totals.iterrows():
    native_charts_ws[f'A{start_row_category+2+idx}'] = row['Category']
    native_charts_ws[f'B{start_row_category+2+idx}'] = float(row['Amount'])

# Create Pie Chart for Spending by Category
pie2 = PieChart()
pie2.title = "Spending by Category"
pie2.title.layout = Layout(manualLayout=ManualLayout(y=0.05, yMode='edge'))  # Position title higher
data2 = Reference(native_charts_ws, min_col=2, min_row=start_row_category+2, max_row=start_row_category+1+len(category_totals))  # Start at +2 to exclude "Amount" header
cats2 = Reference(native_charts_ws, min_col=1, min_row=start_row_category+2, max_row=start_row_category+1+len(category_totals))
pie2.add_data(data2, titles_from_data=False)
pie2.set_categories(cats2)
pie2.legend.position = 'r'  # Position legend to the right
pie2.height = 10
pie2.width = 15  # Increased width to make room for legend on the right
native_charts_ws.add_chart(pie2, f"D{start_row_category}")

# Write data for Spending by Vendor (columns A-B, with 10 rows spacing after previous data+chart pair)
# Category data: title at start_row_category, headers at start_row_category+1, data at start_row_category+2 to start_row_category+2+len(category_totals)-1
# So data ends at start_row_category + 1 + len(category_totals)
category_data_end = start_row_category + 1 + len(category_totals)  # Row where category data ends
start_row_vendor = category_data_end + 10  # 10 rows spacing before next pair
native_charts_ws[f'A{start_row_vendor}'] = 'Spending by Vendor'
native_charts_ws[f'A{start_row_vendor}'].font = Font(bold=True, size=12)
native_charts_ws[f'A{start_row_vendor+1}'] = 'Vendor'
native_charts_ws[f'B{start_row_vendor+1}'] = 'Amount'
native_charts_ws[f'A{start_row_vendor+1}'].font = Font(bold=True)
native_charts_ws[f'B{start_row_vendor+1}'].font = Font(bold=True)
for idx, row in enumerate(vendor_totals.head(15).itertuples(index=False)):  # Limit to top 15 vendors, use enumerate for sequential index
    native_charts_ws[f'A{start_row_vendor+2+idx}'] = row.Vendor
    native_charts_ws[f'B{start_row_vendor+2+idx}'] = float(row.Amount)

# Create Bar Chart for Spending by Vendor (smaller size)
bar = BarChart()
bar.type = "bar"
bar.style = 10
bar.title = "Spending by Vendor"
bar.y_axis.title = 'Amount'
bar.x_axis.title = 'Vendor'
data3 = Reference(native_charts_ws, min_col=2, min_row=start_row_vendor+2, max_row=start_row_vendor+1+min(15, len(vendor_totals)))  # Start at +2 to exclude "Amount" header
cats3 = Reference(native_charts_ws, min_col=1, min_row=start_row_vendor+2, max_row=start_row_vendor+1+min(15, len(vendor_totals)))
bar.add_data(data3, titles_from_data=False)
bar.set_categories(cats3)
bar.height = 10
bar.width = 15
native_charts_ws.add_chart(bar, f"D{start_row_vendor}")

# Adjust column widths
native_charts_ws.column_dimensions['A'].width = 25
native_charts_ws.column_dimensions['B'].width = 15

print("Created native Excel charts in 'Native Charts' sheet")

# Make sure the output directory exists
excel_path.parent.mkdir(parents=True, exist_ok=True)

wb.save(excel_path)
print("Saved Excel file to:", excel_path)
